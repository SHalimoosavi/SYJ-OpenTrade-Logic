"""
SYJ OpenTrade Logic - Report generation (v0.8.0)
===================================================
CSV/Excel exports of classification history and product catalogs, plus a
professional single-classification PDF report showing the full GRI
decision path (the same explainability principle as the rest of this
project -- a report a customs broker could actually use, not just a
data dump).
"""

import csv
import io
from datetime import datetime, timezone
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable


# ---------------------------------------------------------------------------
# CSV / Excel exports
# ---------------------------------------------------------------------------

def classifications_to_csv(rows: List[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["ID", "Product Description", "HTS Code", "Description", "Confidence", "Classified", "Duty Rate", "Date"])
    for r in rows:
        writer.writerow([
            r["id"], r["product_description"], r.get("final_code") or "",
            r.get("final_description") or "", r.get("confidence") or "",
            "Yes" if r["is_classified"] else "No", r.get("duty_rate") or "", r["created_at"],
        ])
    return buf.getvalue().encode("utf-8")


def classifications_to_excel(rows: List[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classifications"

    headers = ["ID", "Product Description", "HTS Code", "Description", "Confidence", "Classified", "Duty Rate", "Date"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")

    for r in rows:
        ws.append([
            r["id"], r["product_description"], r.get("final_code") or "",
            r.get("final_description") or "", r.get("confidence") or "",
            "Yes" if r["is_classified"] else "No", r.get("duty_rate") or "", str(r["created_at"]),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def products_to_csv(rows: List[dict]) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["SKU", "Name", "Description", "HTS Code", "Duty Rate", "Created", "Updated"])
    for r in rows:
        writer.writerow([
            r["sku"], r["name"], r.get("description") or "",
            r.get("hts_code") or "", r.get("duty_rate") or "", r["created_at"], r["updated_at"],
        ])
    return buf.getvalue().encode("utf-8")


def products_to_excel(rows: List[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    headers = ["SKU", "Name", "Description", "HTS Code", "Duty Rate", "Created", "Updated"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="6D28D9", end_color="6D28D9", fill_type="solid")

    for r in rows:
        ws.append([
            r["sku"], r["name"], r.get("description") or "",
            r.get("hts_code") or "", r.get("duty_rate") or "", str(r["created_at"]), str(r["updated_at"]),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF classification report
# ---------------------------------------------------------------------------

def classification_to_pdf(result: dict) -> bytes:
    """
    Builds a professional PDF report for one classification result,
    showing the full GRI decision path -- same explainability principle
    as the dashboard UI, just in a shareable document a customs broker or
    auditor could actually use.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch)
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle("SYJTitle", parent=styles["Title"], textColor=colors.HexColor("#4338CA"), fontSize=18)
    heading_style = ParagraphStyle("SYJHeading", parent=styles["Heading2"], textColor=colors.HexColor("#4338CA"), spaceBefore=14)
    body_style = styles["Normal"]
    small_style = ParagraphStyle("SYJSmall", parent=styles["Normal"], fontSize=8, textColor=colors.grey)

    story = []
    story.append(Paragraph("SYJ OpenTrade Logic &mdash; Classification Report", title_style))
    story.append(Paragraph(f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC", small_style))
    story.append(Spacer(1, 12))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 12))

    story.append(Paragraph("Product Description", heading_style))
    story.append(Paragraph(result["product_description"], body_style))

    story.append(Paragraph("Classification Result", heading_style))
    if result.get("is_classified"):
        summary_data = [
            ["HTS Code", result.get("final_code") or "-"],
            ["Description", result.get("final_description") or "-"],
            ["Confidence", f"{round((result.get('confidence') or 0) * 100)}%"],
            ["Duty Rate", result.get("duty_rate") or "-"],
        ]
        t = Table(summary_data, colWidths=[1.5 * inch, 4.5 * inch])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F5F3FF")),
            ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#4338CA")),
            ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(t)
    else:
        story.append(Paragraph(f"UNRESOLVED: {result.get('unresolved_reason') or 'No reason provided.'}", body_style))

    if result.get("decision_path"):
        story.append(Paragraph("Decision Path", heading_style))
        for step in result["decision_path"]:
            story.append(Paragraph(
                f"<b>[{step['rule_applied']}] {step['node_code']}</b> &mdash; {step['node_description']}",
                body_style,
            ))
            story.append(Paragraph(f"<i>{step['reasoning']}</i> (score: {step['score']:.2f})", small_style))
            story.append(Spacer(1, 6))

    if result.get("alternatives"):
        story.append(Paragraph("Alternatives Considered", heading_style))
        for alt in result["alternatives"]:
            story.append(Paragraph(
                f"{alt['code']} &mdash; {alt['description']} (confidence: {round(alt['confidence'] * 100)}%)",
                small_style,
            ))

    if result.get("related_rulings"):
        story.append(Paragraph("Related CBP CROSS Rulings", heading_style))
        for ruling in result["related_rulings"]:
            story.append(Paragraph(f"<b>{ruling['id']}</b> ({ruling['date']}) &mdash; {ruling['title']}", small_style))

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", color=colors.HexColor("#E5E7EB")))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "This report is generated by a deterministic rules engine (General Rules of Interpretation) for "
        "informational purposes and does not constitute binding customs advice. Verify with a licensed "
        "customs broker before relying on it for an import filing.",
        small_style,
    ))

    doc.build(story)
    return buf.getvalue()
