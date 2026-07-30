import datetime
import io
import os
import sys
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from server_fastapi.reports import (  # noqa: E402
    classifications_to_csv,
    classifications_to_excel,
    products_to_csv,
    products_to_excel,
    classification_to_pdf,
)

SAMPLE_CLASSIFICATIONS = [
    {"id": 1, "product_description": "cordless electric drill", "final_code": "8467.21.00.10",
     "final_description": "Battery powered", "confidence": 0.97, "is_classified": True,
     "duty_rate": "Free", "created_at": datetime.datetime(2026, 7, 26, 10, 0)},
    {"id": 2, "product_description": "xyzzy nonsense", "final_code": None,
     "final_description": None, "confidence": None, "is_classified": False,
     "duty_rate": None, "created_at": datetime.datetime(2026, 7, 26, 10, 5)},
]

SAMPLE_PRODUCTS = [
    {"sku": "DRILL-001", "name": "Cordless Drill", "description": None, "hts_code": "8467.21.00.10",
     "duty_rate": "Free", "created_at": datetime.datetime(2026, 7, 26), "updated_at": datetime.datetime(2026, 7, 26)},
]

SAMPLE_CLASSIFICATION_RESULT = {
    "product_description": "cordless electric drill",
    "final_code": "8467.21.00.10",
    "final_description": "Battery powered",
    "confidence": 0.97,
    "is_classified": True,
    "duty_rate": "Free",
    "decision_path": [
        {"rule_applied": "GRI 1", "node_code": "8467", "node_description": "Tools for working in the hand",
         "reasoning": "Heading 8467 scored highest.", "score": 0.97},
    ],
    "alternatives": [],
    "related_rulings": [],
}


class TestCSVReports(unittest.TestCase):
    def test_classifications_csv_has_header_and_rows(self):
        result = classifications_to_csv(SAMPLE_CLASSIFICATIONS).decode()
        self.assertIn("Product Description", result)
        self.assertIn("cordless electric drill", result)
        self.assertIn("8467.21.00.10", result)

    def test_classifications_csv_handles_unresolved_rows(self):
        result = classifications_to_csv(SAMPLE_CLASSIFICATIONS).decode()
        self.assertIn("xyzzy nonsense", result)
        self.assertIn(",No,", result)

    def test_products_csv_has_header_and_rows(self):
        result = products_to_csv(SAMPLE_PRODUCTS).decode()
        self.assertIn("SKU", result)
        self.assertIn("DRILL-001", result)


class TestExcelReports(unittest.TestCase):
    def test_classifications_excel_is_valid_and_readable(self):
        import openpyxl
        xlsx_bytes = classifications_to_excel(SAMPLE_CLASSIFICATIONS)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        self.assertEqual(ws.title, "Classifications")
        self.assertEqual(ws["A1"].value, "ID")
        self.assertEqual(ws["A2"].value, 1)
        self.assertEqual(ws["B2"].value, "cordless electric drill")

    def test_products_excel_is_valid_and_readable(self):
        import openpyxl
        xlsx_bytes = products_to_excel(SAMPLE_PRODUCTS)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        self.assertEqual(ws.title, "Products")
        self.assertEqual(ws["A2"].value, "DRILL-001")

    def test_empty_rows_do_not_crash(self):
        import openpyxl
        xlsx_bytes = classifications_to_excel([])
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        self.assertEqual(wb.active["A1"].value, "ID")


class TestPDFReports(unittest.TestCase):
    def test_pdf_has_valid_header(self):
        pdf_bytes = classification_to_pdf(SAMPLE_CLASSIFICATION_RESULT)
        self.assertEqual(pdf_bytes[:4], b"%PDF")

    def test_pdf_contains_expected_content(self):
        from pypdf import PdfReader
        pdf_bytes = classification_to_pdf(SAMPLE_CLASSIFICATION_RESULT)
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = reader.pages[0].extract_text()
        self.assertIn("8467.21.00.10", text)
        self.assertIn("Classification Report", text)
        self.assertIn("cordless electric drill", text)

    def test_pdf_for_unresolved_classification_does_not_crash(self):
        unresolved = {
            "product_description": "xyzzy nonsense",
            "final_code": None,
            "is_classified": False,
            "unresolved_reason": "No heading matched.",
            "decision_path": [],
            "alternatives": [],
            "related_rulings": [],
        }
        pdf_bytes = classification_to_pdf(unresolved)
        self.assertEqual(pdf_bytes[:4], b"%PDF")

        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(pdf_bytes))
        text = reader.pages[0].extract_text()
        self.assertIn("UNRESOLVED", text)


if __name__ == "__main__":
    unittest.main()
